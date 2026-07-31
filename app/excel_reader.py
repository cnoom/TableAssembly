"""Excel 解析器:把 .xlsx 读成 TableSchema。

表规范:
- A 列是「定义列」,标识该行的类型。B 列起为字段区,B 列为主键。
- 行类型标记(写在 A 列):
    ##      字段类型行
    #name   字段名行(代码标识符)
    #desc   字段说明行(注释)
    #cs     归属行: both/client/server
- 其它非空 A 列 -> 警告并忽略该行
- 空 A 列 -> 数据行

字段类型(## 行)语法:
    int  float  bool  string        基础类型
    int[] float[] bool[] string[]   数组(默认逗号分隔)
    int[sep=|] string[sep=;] ...    数组 + 自定义分隔符(支持单/多字符)
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from . import rules as R
from . import schema as S
from .schema import (
    BASE_TYPES,
    ARRAY_OF,
    FieldDef,
    DataRow,
    TableSchema,
    VALID_SIDES,
)

# A 列行类型标记
MARK_TYPE = "##"
MARK_NAME = "#name"
MARK_DESC = "#desc"
MARK_CS = "#cs"
# #rule 可带 sep 配置:#rule 或 #rule[sep=...];用前缀匹配识别
MARK_RULE_PREFIX = "#rule"

# #rule[sep=...] 解析正则:group1 = sep 内容
_RULE_MARK_RE = re.compile(r"^#rule(?:\[\s*sep\s*=\s*(.*?)\s*\])?$")

_KNOWN_MARKS = (MARK_TYPE, MARK_NAME, MARK_DESC, MARK_CS)

# 解析类型声明,如 int / int[] / int[sep=|] / string[sep=##]
#   group1: 基础类型名(int/float/bool/string)
#   group2: [] 整体(含可选 sep=...),非数组时为 None
_TYPE_RE = re.compile(r"^\s*(int|float|bool|string)(\[\s*(?:sep\s*=\s*(.*?))?\s*\])?\s*$", re.IGNORECASE)


# ---------------- 类型声明解析 ----------------

def parse_type_decl(text: str) -> tuple[int, str] | None:
    """解析类型声明文本,返回 (type_code, sep)。无法解析返回 None。

    sep: 数组类型的分隔符(空字符串表示默认逗号);非数组返回 ""。
    """
    if text is None:
        return None
    m = _TYPE_RE.match(str(text).strip())
    if not m:
        return None
    base = m.group(1).lower()
    arr = m.group(2)
    base_code = BASE_TYPES.get(base)
    if base_code is None:
        return None
    if arr is None:
        # 非数组
        return base_code, ""
    # 数组:group(3) 为 sep 内容,未指定则默认逗号
    sep_raw = m.group(3)
    sep = sep_raw if sep_raw is not None else ","
    return ARRAY_OF[base_code], sep


# ---------------- 单元格值转换 ----------------

def _parse_scalar(text: Any, type_code: int) -> tuple[bool, Any, str]:
    """把单元格文本转成标量。返回 (ok, value, error_msg)。"""
    if text is None:
        text = ""
    s = str(text).strip()
    if type_code == S.T_INT:
        if s == "":
            return False, None, "空值,无法转为 int"
        try:
            # 避免浮点字符串被 int() 拒绝;统一先转 float 再判断
            f = float(s)
            if not f.is_integer():
                return False, None, f"不是整数: {s!r}"
            return True, int(f), ""
        except ValueError:
            return False, None, f"不是有效 int: {s!r}"
    if type_code == S.T_FLOAT:
        if s == "":
            return False, None, "空值,无法转为 float"
        try:
            return True, float(s), ""
        except ValueError:
            return False, None, f"不是有效 float: {s!r}"
    if type_code == S.T_BOOL:
        low = s.lower()
        if low in ("1", "true", "yes", "y", "是"):
            return True, True, ""
        if low in ("0", "false", "no", "n", "否", ""):
            return True, False, ""
        return False, None, f"不是有效 bool: {s!r}"
    if type_code == S.T_STRING:
        return True, s, ""
    return False, None, f"未知标量类型 code={type_code}"


def _parse_cell(text: Any, field: FieldDef) -> tuple[bool, Any, str]:
    """解析一个单元格(标量或数组)。返回 (ok, value, error_msg)。"""
    if not field.is_array:
        return _parse_scalar(text, field.type_code)
    # 数组
    if text is None:
        text = ""
    s = str(text).strip()
    if s == "":
        return True, [], ""  # 空数组视为合法(检查阶段会给警告)
    elem_code = field.element_type
    parts = s.split(field.sep)
    out: list[Any] = []
    for p in parts:
        ok, v, msg = _parse_scalar(p.strip(), elem_code)
        if not ok:
            return False, None, f"数组元素错误({msg})"
        out.append(v)
    return True, out, ""


# ---------------- 表读取 ----------------

def read_table(file_path: str | Path) -> TableSchema:
    """读取单个 .xlsx 文件为 TableSchema。

    单元格值转换失败不会抛异常,而是把错误塞进 schema.parse_errors,
    并把对应单元格值留为 None,由 checker 统一汇报。
    """
    p = Path(file_path)
    schema = TableSchema(name=p.stem, file_name=p.name)

    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    # 读取第一个工作表(表名通常即文件名,但允许任意)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    type_row = None
    name_row = None
    desc_row = None
    cs_row = None
    # 已收集的定义行的行号,用于诊断
    seen_marks: dict[str, int] = {}
    data_rows_raw: list[tuple[int, tuple]] = []
    rule_rows: list[tuple[int, tuple]] = []  # [(excel_row, row), ...] —— #rule 行(可能多行)
    rule_sep_locked: bool = False  # 是否已通过 #rule[sep=...] 显式锁定 sep

    for idx, row in enumerate(rows_iter):
        excel_row = idx + 1
        if not row:
            continue
        a = row[0]
        a_str = "" if a is None else str(a).strip()
        if a_str == "":
            # 数据行
            data_rows_raw.append((excel_row, row))
            continue
        # #rule 行(可带 [sep=...])
        rule_m = _RULE_MARK_RE.match(a_str) if a_str.startswith(MARK_RULE_PREFIX) else None
        if rule_m is not None:
            # 解析 sep 配置
            sep_raw = rule_m.group(1)
            if sep_raw is not None:
                if sep_raw == "":
                    schema.parse_errors.append(f"行 {excel_row}: #rule[sep=] 的 sep 为空")
                elif sep_raw == ",":
                    schema.parse_errors.append(
                        f"行 {excel_row}: #rule 的 sep 不能是 ','(与规则参数分隔符冲突)"
                    )
                elif not rule_sep_locked:
                    # 首次出现有效 sep 配置,采纳并锁定
                    schema.rule_sep = sep_raw
                    rule_sep_locked = True
                elif sep_raw != schema.rule_sep:
                    schema.parse_errors.append(
                        f"行 {excel_row}: #rule 的 sep 配置 {sep_raw!r} 与首次 {schema.rule_sep!r} 不一致,已忽略"
                    )
            rule_rows.append((excel_row, row))
            continue
        if a_str in _KNOWN_MARKS:
            if a_str in seen_marks:
                schema.parse_errors.append(
                    f"行 {excel_row}: 重复的 {a_str} 定义行(首次出现于行 {seen_marks[a_str]})"
                )
            seen_marks[a_str] = excel_row
            if a_str == MARK_TYPE:
                type_row = row
            elif a_str == MARK_NAME:
                name_row = row
            elif a_str == MARK_DESC:
                desc_row = row
            else:  # MARK_CS
                cs_row = row
        else:
            schema.parse_errors.append(
                f"行 {excel_row}: A 列标记 {a_str!r} 未识别,该行被忽略"
            )

    wb.close()

    # —— 必须有 ## 和 #name ——
    if type_row is None:
        schema.parse_errors.append("缺少 ## 字段类型行,无法解析表结构")
        return schema
    if name_row is None:
        schema.parse_errors.append("缺少 #name 字段名行,无法解析表结构")
        return schema

    # 字段区从 B 列(index=1)开始
    type_cells = list(type_row[1:])
    name_cells = list(name_row[1:]) if name_row else []
    desc_cells = list(desc_row[1:]) if desc_row else []
    cs_cells = list(cs_row[1:]) if cs_row else []

    # 列数以 type 行为准
    n = len(type_cells)
    if n == 0:
        schema.parse_errors.append("## 行没有任何字段(B 列起为空)")
        return schema

    # 定义行列数对齐检查
    for mark, cells, excel_row in (
        (MARK_NAME, name_cells, seen_marks.get(MARK_NAME)),
        (MARK_DESC, desc_cells, seen_marks.get(MARK_DESC)),
        (MARK_CS, cs_cells, seen_marks.get(MARK_CS)),
    ):
        if cells and len(cells) != n:
            schema.parse_errors.append(
                f"{mark} 行(行 {excel_row})列数 {len(cells)} 与 ## 行列数 {n} 不一致"
            )

    # —— 构造字段 ——
    for i in range(n):
        type_text = type_cells[i]
        parsed = parse_type_decl(type_text)
        if parsed is None:
            schema.parse_errors.append(
                f"字段 #{i + 1}(列 {chr(ord('A') + i + 1)}): 类型声明 {type_text!r} 无法解析"
            )
            continue
        type_code, sep = parsed
        name = str(name_cells[i]).strip() if i < len(name_cells) and name_cells[i] else f"field{i}"
        desc = str(desc_cells[i]).strip() if i < len(desc_cells) and desc_cells[i] else ""
        side = str(cs_cells[i]).strip().lower() if i < len(cs_cells) and cs_cells[i] else "both"
        if side not in VALID_SIDES:
            schema.parse_errors.append(
                f"字段 {name}(列 {chr(ord('A') + i + 1)}): 归属 {side!r} 不合法,应为 both/client/server,临时按 both 处理"
            )
            side = "both"
        schema.fields.append(
            FieldDef(name=name, type_code=type_code, sep=sep, desc=desc, side=side)
        )

    # 缺失字段定义的列(类型解析失败)在 fields 中跳过了,后续按 fields 索引读取数据
    if not schema.fields:
        schema.parse_errors.append("没有任何字段成功解析")
        return schema

    # —— 解析数据行 ——
    # 注意:数据行列数与字段数可能不一致(因为有的列类型解析失败被跳过)
    # 这里按 fields 的"原始列下标"读取。fields 的顺序与 type 行一致,
    # 但跳过了 parse 失败的列 —— 为稳健起见,建立 fields 原始列下标映射。
    field_indices = _build_field_indices(type_cells)

    # —— 填充 #rule 规则(多行累加、稀疏、去重)——
    _apply_rule_rows(schema, rule_rows, field_indices)

    for excel_row, row in data_rows_raw:
        values: list[Any] = []
        key: Any = None
        skip_row = False
        for fi, fdef in enumerate(schema.fields):
            col_idx = field_indices.get(fi)
            if col_idx is None or col_idx >= len(row):
                values.append(None if fdef.is_array else None)
                continue
            cell = row[col_idx]
            ok, v, msg = _parse_cell(cell, fdef)
            if not ok:
                # 不在 reader 里堆诊断(checker 统一管),这里只记录无法解析,
                # 把值置 None,checker 会按字段类型重新校验并给出精确行列。
                values.append(None)
                continue
            values.append(v)
        # 主键 = B 列(field 0)的原始文本(未转换),由 checker 校验唯一/非空
        key_col = field_indices.get(0)
        key_raw = None
        if key_col is not None and key_col < len(row):
            key_raw = row[key_col]
            key_raw = "" if key_raw is None else str(key_raw).strip()
        schema.rows.append(DataRow(key=key_raw, values=values, row_index=excel_row))

    return schema


def _build_field_indices(type_cells: list[Any]) -> dict[int, int]:
    """建立 schema.fields[i] -> 原始列下标(在 row tuple 中,B 列=1)的映射。

    fields 的构造顺序就是遍历 type_cells 的顺序,且只保留解析成功的列,
    所以这里需要重新匹配一次:对每个成功解析的位置,给 fields 的下一个。
    """
    mapping: dict[int, int] = {}
    fi = 0
    for i, t in enumerate(type_cells):
        if parse_type_decl(t) is not None:
            mapping[fi] = i + 1  # +1 因为 row[0] 是 A 列
            fi += 1
    return mapping


def _apply_rule_rows(
    schema: TableSchema,
    rule_rows: list[tuple[int, tuple]],
    field_indices: dict[int, int],
) -> None:
    """把 #rule 行的单元格规则累加进对应 FieldDef.rules。

    - 多行累加;空单元格跳过(稀疏)
    - 按 (name, tuple(params)) 去重,保留首次顺序
    - 单元格用 schema.rule_sep 切分多规则;每段 parse_rule 解析
    """
    # 建立 列下标(B=1) -> field 索引的反查表
    col_to_field: dict[int, int] = {col: fi for fi, col in field_indices.items()}
    for excel_row, row in rule_rows:
        # 字段区从 B 列(index=1)开始;对齐到 fields 的原始列下标
        cells = list(row[1:]) if len(row) > 1 else []
        for col_idx in range(1, len(cells) + 1):
            cell = cells[col_idx - 1]
            if cell is None or str(cell).strip() == "":
                continue  # 稀疏:空单元格跳过
            fi = col_to_field.get(col_idx)
            if fi is None or fi >= len(schema.fields):
                continue  # 该列无对应字段(类型解析失败被跳过)
            fdef = schema.fields[fi]
            text = str(cell).strip()
            # 第1层切分:rule_sep 切多规则
            segments = text.split(schema.rule_sep)
            for seg in segments:
                seg = seg.strip()
                if seg == "":
                    continue  # 残留空段跳过
                fr = R.parse_rule(seg)
                if fr is None:
                    continue
                # 去重:按 (name, tuple(params))
                key = (fr.name, tuple(fr.params))
                if any((r.name, tuple(r.params)) == key for r in fdef.rules):
                    continue
                fdef.rules.append(fr)


def scan_directory(dir_path: str | Path) -> list[TableSchema]:
    """扫描目录下所有 .xlsx(忽略 ~$ 临时文件),返回 TableSchema 列表。"""
    p = Path(dir_path)
    if not p.is_dir():
        return []
    out: list[TableSchema] = []
    for f in sorted(p.glob("*.xlsx")):
        if f.name.startswith("~$"):
            continue
        try:
            out.append(read_table(f))
        except Exception as e:  # noqa: BLE001
            s = TableSchema(name=f.stem, file_name=f.name)
            s.parse_errors.append(f"读取文件失败: {e!r}")
            out.append(s)
    return out
